from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
from helpers import char_range, Employee, create_generic_import
import win32com.client as win32
import os

# Opens the Novatime export, saves it as xlsx
fname = os.path.abspath('TWKPR.XLS')
excel = win32.gencache.EnsureDispatch('Excel.Application')
toSave = excel.Workbooks.Open(fname)
toSave.SaveAs(fname+"X", FileFormat=51)
toSave.Close()
excel.Application.Quit()

# Opens the new xlsx file for manipulation
wb = load_workbook(filename="TWKPR.xlsx", read_only=True)
sheet = wb.active
center_aligned_text = Alignment(horizontal="center")


def newTimecard(id, type, hours):
    """Creates a new Employee dataclass instance and stores the employee id and either the regular or overtime hours associated with it in the first cell found with that employee id"""
    emp = Employee(id=id)
    if type == 'reg':
        emp.reg = hours
    if type == 'ot1':
        emp.ot1 = hours
    return emp


def convertPPNT():
    """Iterates over a Novatime export file, converting the data into an Employee dataclass."""
    # List for the Employee data to be stored
    data = []
    # Iterates over the Novatime export
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True):
        # If there is already data in the data list
        if len(data) > 0:
            # if the employee id in the last index of the data list is the same as the current one
            if data[len(data)-1].id == row[0]:
                # if the current row is a regular time row
                if row[3].lower() == 'reg':
                    # set the value of reg to the value in the current iteration
                    data[len(data)-1].reg = row[4]
                else:
                    # Set the value of ot1 to the value in the current iteration
                    data[len(data)-1].ot1 = row[4]
            else:
                # Create a new time card for the current iteration's employee number
                tc = newTimecard(row[0], row[3].lower(), row[4])
                data.append(tc)
        else:
            # Create a new time card for the current iteration's employee number
            tc = newTimecard(row[0], row[3].lower(), row[4])
            data.append(tc)
    return data


import_data = convertPPNT()
create_generic_import(["Novatime", import_data], 1.165, "Papa Pita")
# Closes workbook
wb.close()
# Deletes xlsx file
os.remove(fname+"X")
