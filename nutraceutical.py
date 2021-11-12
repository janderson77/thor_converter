from openpyxl import load_workbook
from helpers import Employee, collect_sheet_names, create_generic_import

def collect_hours(sheet):
    """Iterates over the sheet data, creating a new Employee class instance, filling it with data and appending it onto the data list"""
    data = []
    for row in sheet.iter_rows(min_row=sheet.min_row, max_row=sheet.max_row, values_only=True):
        if row[1] != None and type(row[1]) == str:
            data.append(Employee(id=row[0]))
        # if row[0] != None and type(row[0]) == int and row[0]>0 and row[1] != None:
        #     data.append(Employee(id=row[0]))
        if row[0] != None and row[2] == None and row[5] != None:
            data[len(data)-1].reg += row[5]
        elif row[3] != None and row[3].lower() == "holiday":
            data[len(data)-1].holiday += float(row[5])
        elif row[0] == None and row[2] == None and row[5] != None:
            data[len(data)-1].reg += float(row[5])
        else:
            continue
        
    for i in data:
        if i.holiday:
            i.reg = i.reg-i.holiday
        if i.reg > 40:
            i.ot1 = round(i.reg-40,2)
            i.reg = 40
        
    return data

def convert_nutra(input):
    """Converts all collected employee data from the client provided sheet into an importable excel file."""

    wb = load_workbook(input, read_only=True)

    sheets = collect_sheet_names(wb,1)

    data = []
    to_export = []

    for s in sheets:
        data.append([s,collect_hours(wb[s])])

    for d in data:
        res = create_generic_import(d, 1.30, "Nutraceutical")
        to_export.append(res)
    
    return to_export[0]