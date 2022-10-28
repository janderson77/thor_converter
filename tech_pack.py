from openpyxl import load_workbook
from helpers import collect_sheet_names

def get_teckpack_data(sheet):
    hazleton_data = []
    hazleton_custom_data = []
    for row in sheet.iter_rows(min_row=sheet.min_row, max_row = sheet.max_row, values_only=True):
        if 'TEMP' in row:

            if 'Custom' in row[1]:
                hazleton_custom_data.append(row[:14])
            else:
                hazleton_data.append(row[:14])

    return(hazleton_data, hazleton_custom_data)

def clear_sheet_data(sheet):
    for row in sheet.iter_rows(min_row=8):
        for cell in row:
            cell.value = None
    return sheet

def insert_sheet_data(sheet, data):
    print("working on this")

wb = load_workbook(filename='Tech Pack Hours 101622.xlsx')
source = wb.active

sheets = collect_sheet_names(wb,1)

gotten_data = get_teckpack_data(wb[sheets[0]])

hazleton = clear_sheet_data(wb.copy_worksheet(source))
hazleton_custom = clear_sheet_data(wb.copy_worksheet(source))

hazleton = insert_sheet_data(hazleton, gotten_data[0])
hazleton_custom = insert_sheet_data(hazleton_custom, gotten_data[1])


