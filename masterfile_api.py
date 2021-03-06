from openpyxl import load_workbook
from helpers import Employee, create_adjustment_import, create_generic_import, collect_sheet_names

def collect_hours(row, number):
    """Collects hours data from the appropriate cells"""
    # Checks if there is data in the cell. If there is data, checks if it is already a float or not. If not, converts to float before returning.
    if type(row[number]) == None:
        return None
    if type(row[number]) == str and len(row[number]) < 1:
        return None
    elif type(row[number]) == float:
        return row[number]
    elif type(row[number]) == str and len(row[number]) > 0:
        print(row[number])
        if row[number] == "#N/A":
            return None
        if "=" in row[number]:
            return None
        elif "." in row[number]:
            return float(row[number])
        else:
            h = int(row[number])
            return float(h)
    elif type(row[number]) == int:
        return float(row[number])
    else:
        return row[number]


def collect_employee_data(sheet, columns, s_name="None"):
    """Iterates over the sheet data, creating a new Employee class instance, filling it with data and appending it onto the data list"""
    data = []
    for row in sheet.iter_rows(min_row=sheet.min_row, max_row=sheet.max_row, values_only=True):
        if row[0] != None and type(row[0]) != str and row[0] <= 1:
            continue
        if row[0] != None and type(row[0]) != str:
            tc = Employee(id=row[0])
            reg1 = collect_hours(row, columns['reg1'])
            reg2 = collect_hours(row, columns['reg2'])
            ot1 = collect_hours(row, columns['ot1'])
            ot2 = collect_hours(row, columns['ot2'])
            salary = collect_hours(row, columns['salary'])
            bonus = collect_hours(row, columns['bonus'])
            commission = collect_hours(row, columns['commission'])
            expenses = collect_hours(row, columns['expenses'])

            # Adds regular hours to time card, or sets the value to None
            if reg1 != None and reg2 != None:
                tc.reg = round(reg1 + reg2, 2)
            elif reg1 != None and reg2 == None:
                tc.reg = round(reg1, 2)
            elif reg1 == None and reg2 != None:
                tc.reg = round(reg2, 2)
            else:
                tc.reg = None

            # Adds overtime hours to time card, or sets the value to None
            if ot1 != None and ot2 != None:
                tc.ot1 = round(ot1 + ot2, 2)
            elif ot1 != None and ot2 == None:
                tc.ot1 = round(ot1, 2)
            elif ot1 == None and ot2 != None:
                tc.ot1 = round(ot2, 2)
            elif ot1 == None and ot2 == None and tc.reg != None:
                tc.ot1 = 0
            else:
                tc.ot1 = None

            # Adds salary, bonus, commision and expense data, or leaves values as 0 or None
            if salary != None:
                tc.salary = round(salary, 2)
            if bonus != None:
                tc.bonus = round(bonus, 2)
            if commission != None:
                tc.commission = round(commission, 2)
            if expenses != None:
                if s_name == "DSD Managers":
                    tc.expenses = round(expenses, 2) - 18
                    tc.adjustment = 18
                else:
                    tc.expenses = round(expenses, 2)

            data.append(tc)
    return data


def find_data_row(sheet):
    """Finds the row where the column names live"""
    label_row = 1
    for row in sheet.iter_rows(min_row=sheet.min_row, max_row=sheet.max_row, values_only=True):
        if row[0] == "Emp #":
            return label_row
        else:
            label_row += 1


def find_data_column(sheet, row):
    """Finds the column numbers for each required column"""
    columns = {
        "reg1": 6,
        "reg2": 6,
        "ot1": 7,
        "ot2": 7,
        "salary": None,
        "bonus": None,
        "commission": None,
        "expenses": 23

    }

    data_row = None

    for row in sheet.iter_rows(min_row=row, max_row=row, values_only=True):
        data_row = row

    for i, v in enumerate(data_row[:23]):
        if v == None:
            continue
        if "Reg Hrs 1 Week" in v:
            columns['reg1'] = i
            columns['reg2'] = (i+1)

        if "OT Hrs Week 1" in v or "OT Hrs week1" in v or "OT Hrs week 1" in v:
            columns['ot1'] = i
            columns['ot2'] = i+1

        if "Salary Pay" in v:
            columns['salary'] = i
        if "Bonus Pay" in v:
            columns['bonus'] = i
        if "Commission Pay" in v:
            columns['commission'] = i
        if "Expenses-" in v or "Expenses -" in v:
            columns['expenses'] = i

    return columns


def convert_masterfile(input):
    """
    Converts all collected employee data from the Masterfile into importable excel files. DSD Managers is treated uniquely as they have cell phone adjustments.
    """
    wb = load_workbook(input, read_only=True)

    sheets = collect_sheet_names(wb,6)
    data = []
    special_data = []
    to_export = []

    for s in sheets:
        row = find_data_row(wb[s])
        columns = find_data_column(wb[s], row)
        if s == "DSD Managers":
            special_data.append([s, collect_employee_data(wb[s], columns, s)])
        else:
            data.append([s, collect_employee_data(wb[s], columns, s)])

    for d in data:
        res = create_generic_import(d, 1.165, "Papa Pita")
        to_export.append(res)

    if len(special_data) > 0:
        for d in special_data:
            gen_res = create_generic_import(d, 1.165, "Papa Pita")
            # adj_res = create_adjustment_import(d, 48, "Papa Pita")
            to_export.append(gen_res)
            # to_export.append(adj_res)

    return to_export
