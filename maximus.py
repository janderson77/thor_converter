from dataclasses import dataclass
from openpyxl import load_workbook
import datetime

test = load_workbook("maximus_test_file.xlsx", read_only=True)

@dataclass
class MaximusTC:
    name: str = None
    paycode: str = None
    line_item_id: str = None
    unit_pay: float = None
    unit_bill: float = None
    hours: float = None
    adjustment_pay: float = None
    adjustment_bill: float = None
    weekend_date: datetime = None

def get_weekend_date(date_string, paycode):
    starting = ""
    if paycode == "sick":
        starting = date_string.find("Pay") + 4
    elif paycode == "covid-19":
        starting = date_string.find("(")-9

    counter = 0
    new_date_string = ""
    while counter <= 7:
        new_date_string = new_date_string + date_string[starting+counter]
        counter+=1
    
    if "-" in new_date_string:
        new_date = datetime.datetime.strptime(new_date_string, '%m/%d/%y').strftime("%m-%d-%y")
    else:
        new_date = datetime.datetime.strptime(new_date_string, '%m/%d/%y')
    weekend_date = new_date + datetime.timedelta(days=6-datetime.datetime.weekday(new_date))
    return weekend_date

def collect_maximux_data(sheet):
    ws = sheet[sheet.sheetnames[0]]
    tcdata = []

    for row in ws.iter_rows(min_row=ws.min_row, max_row=100, values_only=True):
        if row[0] == None:
            break
        if row[0].lower() == "id":
            continue

        employee = MaximusTC()

        employee_name_end_index = row[1].find(" -")
        employee.name = row[1][:employee_name_end_index]
        employee.line_item_id = row[0]

        if "sick" in row[1].lower():
            employee.paycode = "sick"

            # Move up in scope when all paycodes are working
            # It takes a paycode as an argument
            employee.weekend_date = get_weekend_date(row[1], employee.paycode)
            starting = row[1].find("(")
            sick_hours = ""
            for index, value in enumerate(row[1], start = starting+1):
                if row[1][index] == " ":
                    employee.hours = float(sick_hours)
                    break
                if row[1][index] != "(":
                    sick_hours = sick_hours + row[1][index]
            tcdata.append(employee)

        elif "covid" in row[1].lower():
            employee.paycode = "covid-19"
            employee.weekend_date = get_weekend_date(row[1], employee.paycode)
            totalHours = None
            ending = row[1].lower().find(" hours")-1
            starting = row[1].find("(")+1
            if starting != ending:
                totalHours = [row[1][starting], row[1][ending]]
                totalHours = float("".join(totalHours))
            else:
                totalHours = float(row[1][starting])
            employee.hours = totalHours
            tcdata.append(employee)

        elif "bonus" in row[1].lower():
            employee.weekend_date = row[2] + datetime.timedelta(days=6-datetime.datetime.weekday(row[2]))
            employee.paycode = "bonus"
            employee.unit_pay = float(row[4])
            employee.unit_bill = float(row[3])
            tcdata.append(employee)

        else:
            # If none of the above, must be an adjustment.
            employee.weekend_date = row[2] + datetime.timedelta(days=6-datetime.datetime.weekday(row[2]))
            if "background" in row[1].lower():
                employee.paycode = "114"
                employee.adjustment_bill = float(row[3])
                tcdata.append(employee)
                continue
            elif "internet" in row[1].lower():
                employee.paycode = "151"
            elif "monitor" in row[1].lower():
                employee.paycode = "27"
            else:
                employee.paycode = "ERROR"
            employee.adjustment_pay = float(row[4])
            employee.adjustment_bill = float(row[3])
            tcdata.append(employee)
    return tcdata
    

test_run = collect_maximux_data(test)
# for i in test_run:
#     print(i)

# Usefule for creating the import
# we_date = datetime.datetime.strptime(datetime.datetime.fromisoformat(str(row[2])).strftime("%m/%d/%y"), '%m/%d/%y')