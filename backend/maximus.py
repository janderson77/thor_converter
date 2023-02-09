from dataclasses import dataclass
from openpyxl import load_workbook
import datetime
from helpers import create_generic_import_with_req_number, create_adjustment_import_with_req_number

@dataclass
class MaximusTC:
    name: str = None
    first: str = None
    last: str = None
    twid: int = None
    paycode: str = None
    line_item_id: str = None
    unit_pay: float = None
    unit_bill: float = None
    hours: float = None
    adjustment_pay: float = None
    adjustment_bill: float = None
    weekend_date: datetime = None

def get_weekend_date(date_string, paycode):
    """
    Finds the date information for the given entry.
    Returns the appropriate week ending date (Always Sunday)
    """
    starting = ""
    if paycode == "sick" or paycode == "covid-19":
        starting = date_string.find("(")-9
        if date_string[starting] == "y" or date_string[starting] == "-":
            starting = starting+2

    counter = 0
    new_date_string = ""
    while counter <= 9:
        if date_string[starting+counter] == " ":
            if counter < 1:
                counter+=1
                continue
            else:
                counter = 10
        else:
            new_date_string = new_date_string + date_string[starting+counter]
            counter+=1

    new_date_string.strip()
    if "-" in new_date_string:
        new_date = datetime.datetime.strptime(new_date_string, '%m/%d/%y').strftime("%m/%d/%y")
    else:
        try:
            new_date = datetime.datetime.strptime(new_date_string, '%m/%d/%y').strftime("%m/%d/%y")
        except ValueError as v:
            if len(v.args) > 0 and v.args[0].startswith('unconverted data remains:'):
                new_date_string = new_date_string[0:len(new_date_string)-1]
                
                new_date = datetime.datetime.strptime(new_date_string, '%m/%d/%y').strftime("%m/%d/%y")
        new_date = datetime.datetime.strptime(new_date_string, '%m/%d/%y').strftime("%m/%d/%y")
    new_date = datetime.datetime.strptime(new_date, '%m/%d/%y')
    weekend_date = new_date + datetime.timedelta(days=6-datetime.datetime.weekday(new_date))
    return weekend_date

def get_twid(employee, report):
    '''
    Gets tempworks employee ID's from an assignment register report and returns them.
    If the function returns None, then they are not on the report.
    '''
    for row in report.iter_rows(min_row=6, max_row=9999, values_only=True):
        if not row[4]:
            break
        if employee.first[:len(employee.first)-1].lower() in row[4].lower() and employee.last[:len(employee.last)-1].lower() in row[4].lower():
            return int(row[31])

def collect_maximux_data(maximus_data, assignment_register):
    tcdata = []
    adjust_data = []

    for row in maximus_data.iter_rows(min_row=maximus_data.min_row, max_row=100, values_only=True):
        if row[0] == None:
            break
        if row[0].lower() == "id":
            continue

        employee = MaximusTC()

        employee_name_end_index = row[1].find(" -")
        employee.name = row[1][:employee_name_end_index]
        if "monitor" in employee.name.lower():
            employee.name = row[1][employee_name_end_index+3:]
        first_last_split = employee.name.find(" ")
        employee.first = employee.name[:first_last_split]
        employee.last = employee.name[first_last_split:]
        employee.line_item_id = row[0]
        employee.twid = get_twid(employee, assignment_register)

        if "sick" in row[1].lower():
            employee.paycode = "sick"

            # Move up in scope when all paycodes are working
            # It takes a paycode as an argument
            employee.weekend_date = get_weekend_date(row[1], employee.paycode)
            starting = row[1].find("(")
            sick_hours = ""
            for index, value in enumerate(row[1], start = starting+1):
                if row[1][index] == " ":
                    employee.hours = float(sick_hours)
                    break
                if row[1][index] != "(":
                    sick_hours = sick_hours + row[1][index]
            tcdata.append(employee)

        elif "covid" in row[1].lower():
            employee.paycode = "covid-19"
            employee.weekend_date = get_weekend_date(row[1], employee.paycode)
            totalHours = None
            ending = row[1].lower().find(" hours")-1
            starting = row[1].find("(")+1
            if starting != ending:
                totalHours = [row[1][starting], row[1][ending]]
                totalHours = float("".join(totalHours))
            else:
                totalHours = float(row[1][starting])
            employee.hours = totalHours
            tcdata.append(employee)

        elif "bonus" in row[1].lower():
            employee.weekend_date = row[2] + datetime.timedelta(days=6-datetime.datetime.weekday(row[2]))
            employee.paycode = "bonus"
            employee.unit_pay = float(row[4])
            employee.unit_bill = float(row[3])
            tcdata.append(employee)

        else:
            # If none of the above, must be an adjustment.
            employee.weekend_date = row[2] + datetime.timedelta(days=6-datetime.datetime.weekday(row[2]))
            if "background" in row[1].lower():
                employee.paycode = "114"
                employee.adjustment_bill = float(row[3])
                adjust_data.append(employee)
                continue
            elif "internet" in row[1].lower():
                employee.paycode = "151"
            elif "monitor" in row[1].lower():
                employee.paycode = "27"
            elif "office" in row[1].lower():
                employee.paycode = "46"
            else:
                employee.paycode = "ERROR"
            employee.adjustment_pay = float(row[4])
            employee.adjustment_bill = float(row[3])
            adjust_data.append(employee)
    return [tcdata, adjust_data]
    

def convert_maximus(maximus_data, assignment_register):
    maximus_wb = load_workbook(maximus_data, read_only=True)
    maximus_ws = maximus_wb[maximus_wb.sheetnames[0]]
    areg_wb = load_workbook(assignment_register, read_only=True)
    areg_ws = areg_wb[areg_wb.sheetnames[0]]

    data = []

    data.append(collect_maximux_data(maximus_ws, areg_ws))
    
    gen_import = create_generic_import_with_req_number(data[0], "Maximus")
    adjust_import = create_adjustment_import_with_req_number(data[0][1], "Maximus")

    return [gen_import, adjust_import]