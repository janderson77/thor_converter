from openpyxl import load_workbook
import csv

csvfilename = "Premier_PapaPitaConsolidatedInv.csv"
xlsxfilename = "20230418 Porsche Remittance.xlsx"


def reconcile_vms(csvfilename, xlsxfilename):
    csvfields = []
    csvrows = []
    xlsxfields = []
    xlsxrows = []

    wb = load_workbook(xlsxfilename, read_only=True)
    ws = wb.active

    # Opens csv file
    with open(csvfilename, "r") as csvfile:
        # Creates csvreader object with csv file contents
        csvreader = csv.reader(csvfile)

        # appends all rows to the csv rows list
        for row in csvreader:
            csvrows.append(row)

    # extracts fields and appends them to csvfields list
    for item in csvrows[3]:
        csvfields.append(item)
    # removes all but the data below the column headers
    csvrows = csvrows[4:]

    # appends all rows to the xlsxrows list
    for row in ws.iter_rows(min_row=ws.min_row, max_row=ws.max_row, values_only=True):
        xlsxrows.append(row)

    # extracts fields and appends them to xlsxfields list
    for item in xlsxrows[0]:
        xlsxfields.append(item)
    # removes all but the data below the column headers
    xlsxrows = xlsxrows[1:]

reconcile_vms(csvfilename, xlsxfilename)