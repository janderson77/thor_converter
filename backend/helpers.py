from dataclasses import dataclass
from openpyxl import Workbook
from openpyxl.styles import Alignment, numbers
import io
from random import randrange
import pyexcel as p

center_aligned_text = Alignment(horizontal='center')


@dataclass
class Employee:
    id: int = None
    reg: float = 0
    ot1: float = 0
    salary: float = 0
    commission: float = 0
    bonus: float = 0
    expenses: float = 0
    adjustment: float = 0
    holiday: float = 0
    vacation: float = None
    miles: float = None


@dataclass
class Columns:
    reg: int = 0
    ot1: int = 0
    salary: int = 0
    bonus: int = 0
    commission: int = 0
    expenses: int = 0
    vacation: int = 0


def collect_sheet_names(input, limit):
    """
    Collects the names of the sheets in a workbook.
    Stops at sheet the given limit as there may be sheets with no data or useful data beyond that
    """
    useful_sheets = input.sheetnames
    return useful_sheets[:limit]


def char_range(c1, c2):
    '''Created the ability to iterate over a range of characters for easier tweaking of cells in a row'''
    for c in range(ord(c1), ord(c2)+1):
        yield chr(c)


def create_adjustment_import(data, adjust_id, customer_name=None):
    '''
    Creates an adjustment import file with the provided data.
    Data must be in a 2d array, with index 0 of the main array being the sheet name, and index 1 the employee timecard data array.
    Enter the appropriate adjustment ID for the adjustment being made.
    Argument 'customer_name' is optional.
    '''
    wb = Workbook()
    new_sheet = wb.active

    for i in char_range('A', 'E'):
        new_sheet[f'{i}1'].alignment = center_aligned_text

    new_sheet['A1'] = 'Customer Name'
    new_sheet['B1'] = 'Adjustment ID'
    new_sheet['C1'] = 'Employee ID'
    new_sheet['D1'] = 'Adjustment Pay'
    new_sheet['E1'] = 'Adjustment Bill'

    # Sets the starting row to be edited as row 2
    sheet_row = 2

    # iterates over the data list
    for e in data[1]:
        # Sets column A to the value of the customer name or leaves blank
        new_sheet.cell(row=sheet_row, column=1).value = customer_name

        # Sets column B to teh value of the adjustment ID
        new_sheet.cell(row=sheet_row, column=2).value = adjust_id

        # Sets column C to the value of employee id
        new_sheet.cell(row=sheet_row, column=3).value = e.id

        # Sets column D and E to the value of the adjustment
        new_sheet.cell(row=sheet_row, column=4).value = e.adjustment
        new_sheet.cell(row=sheet_row, column=5).value = e.adjustment

        sheet_row += 1

    # Saves as a new file
    return wb.save(filename=f'uploads/{customer_name} Adjustment Import File.xlsx')


def create_pbm_import(data):
    '''
    Creates a PBM import with the provided data.
    Data must be in an list, and data must be in a "Employee" dataclass format.
    '''

    sheet = p.Sheet()

    sheet[0, 0] = "Co Code"
    sheet[0, 1] = "Batch ID"
    sheet[0, 2] = "File #"
    sheet[0, 3] = "Reg Hours"
    sheet[0, 4] = "O/T Hours"

    row = 1

    for emp in data:
        sheet[row, 2] = emp.id
        sheet[row, 3] = emp.reg
        sheet[row, 4] = emp.ot1
        row += 1

    s = sheet.save_to_memory(file_type='xls')

    return([s, "PBM Import"])


def create_generic_import(data, markup, customer_name=None):
    '''
    Creates a generic timecard import with the provided data.
    Data must be in a 2d array, with index 0 of the main array being the sheet name, and index 1 the employee timecard data array.
    Markup is requird. Must be a float. Example: 1.165
    Argument 'customer_name' is optional.
    '''

    if customer_name == None:
        customer_name = ""
    wb = Workbook()
    new_sheet = wb.active

    for i in char_range('A', 'G'):
        new_sheet[f'{i}1'].alignment = center_aligned_text

    new_sheet['A1'] = 'Employee Name'
    new_sheet['B1'] = 'Emp #'
    new_sheet['C1'] = 'Cust Name'
    new_sheet['D1'] = 'Pay Rate'
    new_sheet['E1'] = 'Bill Rate'
    new_sheet['F1'] = 'Reg Hrs'
    new_sheet['G1'] = 'OT Hrs'
    new_sheet['H1'] = 'DT Hrs'
    new_sheet['I1'] = 'Paycode'
    new_sheet['J1'] = 'Units'
    new_sheet['K1'] = 'Unit Pay'
    new_sheet['L1'] = 'Unit Bill'
    new_sheet['M1'] = 'Salary Pay'
    new_sheet['N1'] = 'Salary Bill'

    # Sets the starting row to be edited as row 2
    sheet_row = 2

    # iterates over the data list
    for e in data[1]:
        # Sets column I to the value of 'Reg'
        new_sheet.cell(row=sheet_row, column=9).value = 'Reg'

        # Sets column B to the value of employee id
        new_sheet.cell(row=sheet_row, column=2).value = e.id

        # Sets column C to the value of the customer name or None
        new_sheet.cell(row=sheet_row, column=3).value = customer_name

        if customer_name.lower() == 'papa pita bakery' or customer_name.lower() == "papa pita":
            if e.id == 293355:  # Sets payrate to specified if they are a special case employee
                new_sheet.cell(row=sheet_row, column=4).value = 100
                new_sheet.cell(row=sheet_row, column=5).value = 116.5
            # Sets bill rate to 0.00 if a Papa Pita employee works less than 4 hours
            elif e.reg != None and e.reg <= 4.00 and data[0] == 'Novatime':
                new_sheet.cell(row=sheet_row, column=5, value=float(0.00))
                new_sheet.cell(
                    row=sheet_row, column=5).number_format = numbers.FORMAT_NUMBER_00
                new_sheet.cell(row=sheet_row, column=9, value=str('reg agree'))

        # Sets column F to the value of regular hours
        if e.reg:
            new_sheet.cell(row=sheet_row, column=6).value = e.reg

        # Sets column G to the value of OT hours
        if e.ot1:
            new_sheet.cell(row=sheet_row, column=7).value = e.ot1
        elif e.reg and not e.ot1:
            new_sheet.cell(row=sheet_row, column=7).value = 0

        # Sets columns J to 1, K to the value of commission, and L to commission times markup
        if e.commission:
            new_sheet.cell(row=sheet_row, column=10).value = 1
            # If there are expenses, which is usually just a Papa Pita thing, they get added to the unit pay and bill
            if e.expenses > 0:
                expenses_plus_commission = e.expenses + e.commission
                new_sheet.cell(
                    row=sheet_row, column=11).value = expenses_plus_commission
                new_sheet.cell(row=sheet_row, column=12).value = (
                    expenses_plus_commission * markup)
            else:
                new_sheet.cell(row=sheet_row, column=11).value = e.commission
                new_sheet.cell(row=sheet_row, column=12).value = (
                    e.commission * markup)

        # Sets columns M to the value of salary, and N to salary times markup
        if e.salary:
            new_sheet.cell(row=sheet_row, column=13).value = e.salary
            new_sheet.cell(row=sheet_row, column=14).value = (
                e.salary * markup)

        # Increments the sheet_row variable so that the next set of data is put on a new row
        sheet_row += 1

        # Creates a new row for a bonus, Holiday or Vacation timecard, Sets appropriate values and increments sheet_row again.
        if e.bonus:
            new_sheet.cell(row=sheet_row, column=2).value = e.id
            new_sheet.cell(row=sheet_row, column=3).value = customer_name
            new_sheet.cell(row=sheet_row, column=9).value = 'Bonus'
            new_sheet.cell(row=sheet_row, column=10).value = 1
            new_sheet.cell(row=sheet_row, column=11).value = e.bonus
            new_sheet.cell(row=sheet_row, column=12).value = (e.bonus * markup)
            sheet_row += 1

        if e.holiday:
            new_sheet.cell(row=sheet_row, column=2).value = e.id
            new_sheet.cell(row=sheet_row, column=3).value = customer_name
            new_sheet.cell(row=sheet_row, column=9).value = 'Hol'
            new_sheet.cell(row=sheet_row, column=6).value = e.holiday

            sheet_row += 1

        if e.vacation:
            new_sheet.cell(row=sheet_row, column=2).value = e.id
            new_sheet.cell(row=sheet_row, column=3).value = customer_name
            new_sheet.cell(row=sheet_row, column=9).value=str('vac1')
            new_sheet.cell(row=sheet_row, column=10).value = 1
            new_sheet.cell(row=sheet_row, column=11).value = e.vacation
            new_sheet.cell(row=sheet_row, column=12).value = e.vacation*markup
            sheet_row += 1
        if e.miles:
            new_sheet.cell(row=sheet_row, column=2).value = e.id
            new_sheet.cell(row=sheet_row, column=3).value = customer_name
            new_sheet.cell(row=sheet_row, column=9).value = 'reg'
            new_sheet.cell(row=sheet_row, column=10).value = 1
            new_sheet.cell(row=sheet_row, column=11).value = e.miles
            new_sheet.cell(row=sheet_row, column=12).value = (round(e.miles * markup,2))
            sheet_row += 1


    # Saves as a new file
    if customer_name == 'Papa Pita' or customer_name == 'Papa Pita Bakery':
        file = io.BytesIO()
        wb.save(file)
        file.seek(0)
        return([file, f'Papa Pita {data[0]} Import'])
    else:
        file = io.BytesIO()
        wb.save(file)
        file.seek(0)
        return([file, f'{customer_name} Generic Import'])

def create_generic_import_with_req_number(data, client_name = None):
    '''
    Creates a generic timecard import with the provided data.
    Data must be in a 2d array, with index 0 of the main array being the sheet name, and index 1 the employee timecard data array.
    '''

    wb = Workbook()
    new_sheet = wb.active

    for i in char_range('A', 'P'):
        new_sheet[f'{i}1'].alignment = center_aligned_text

    new_sheet['A1'] = 'Employee Name'
    new_sheet['B1'] = 'Emp #'
    new_sheet['C1'] = 'Req Number'
    new_sheet['D1'] = 'Cust Name'
    new_sheet['E1'] = 'WeekendDate'
    new_sheet['F1'] = 'Pay Rate'
    new_sheet['G1'] = 'Bill Rate'
    new_sheet['H1'] = 'Reg Hrs'
    new_sheet['I1'] = 'OT Hrs'
    new_sheet['J1'] = 'DT Hrs'
    new_sheet['K1'] = 'Paycode'
    new_sheet['L1'] = 'Units'
    new_sheet['M1'] = 'Unit Pay'
    new_sheet['N1'] = 'Unit Bill'
    new_sheet['O1'] = 'Salary Pay'
    new_sheet['P1'] = 'Salary Bill'

    # Sets the starting row to be edited as row 2
    sheet_row = 2

    # iterates over the data list
    for e in data[0]:
        if e.adjustment_pay:
            continue

        # Sets column A to the employee's name
        new_sheet.cell(row=sheet_row, column=1).value = e.name

        # Sets column B to the value of employee id
        new_sheet.cell(row=sheet_row, column=2).value = e.twid

        # Sets column C to the value of the req number
        new_sheet.cell(row=sheet_row, column=3).value = e.line_item_id

        # sets column E to the weekend date
        new_sheet.cell(row=sheet_row, column=5).value = e.weekend_date

        # Sets column K to the value of appropriate paycode
        new_sheet.cell(row=sheet_row, column=11).value = e.paycode

        # Sets column H to the value of regular hours
        if e.hours:
            new_sheet.cell(row=sheet_row, column=8).value = e.hours
        
        # Sets column L to 1, M to unit pay and N to unit bill
        if e.unit_pay:
            new_sheet.cell(row=sheet_row, column=12).value = 1
            new_sheet.cell(row=sheet_row, column=13).value = e.unit_pay
            new_sheet.cell(row=sheet_row, column=14).value = e.unit_bill

        # Increments the sheet_row variable so that the next set of data is put on a new row
        sheet_row += 1

    # Saves as a new file
    file = io.BytesIO()
    wb.save(file)
    file.seek(0)
    if client_name != None:
        return([file, f'{client_name} Generic Import With Req Number'])
    else:
        return([file, f'Generic Import With Req Number'])

def create_adjustment_import_with_req_number(data, client_name=None):
    '''
    Creates an adjustment import file with the provided data.
    '''
    wb = Workbook()
    new_sheet = wb.active

    for i in char_range('A', 'G'):
        new_sheet[f'{i}1'].alignment = center_aligned_text

    new_sheet['A1'] = 'Customer Name'
    new_sheet['B1'] = 'Adjustment ID'
    new_sheet['C1'] = 'Employee ID'
    new_sheet['D1'] = 'Req Number'
    new_sheet['E1'] = 'Adjustment Pay'
    new_sheet['F1'] = 'Adjustment Bill'
    new_sheet['G1'] = 'Invoice Text'


    # Sets the starting row to be edited as row 2
    sheet_row = 2

    # iterates over the data list
    for e in data:
        if e.hours or e.unit_pay:
            continue
        # Sets column B to teh value of the adjustment ID
        
        if(e.paycode == "ERROR"):
            new_sheet.cell(row=sheet_row, column=2).value = e.paycode
        else:
            new_sheet.cell(row=sheet_row, column=2).value = int(e.paycode)

        # Sets column C to the value of employee id
        new_sheet.cell(row=sheet_row, column=3).value = e.twid

        # Sets column D to the req number
        new_sheet.cell(row=sheet_row, column=4).value = e.line_item_id

        # Sets column E and F to the value of the adjustment
        new_sheet.cell(row=sheet_row, column=5).value = e.adjustment_pay
        new_sheet.cell(row=sheet_row, column=6).value = e.adjustment_bill

        sheet_row += 1

    # Saves as a new file and returns the file
    file = io.BytesIO()
    wb.save(file)
    file.seek(0)
    if client_name != None:
        return([file, f'{client_name} Adjustment Import With Req Number'])
    else:
        return([file, f'Adjustment Import With Req Number'])


def getRandomPhrase():
    phraseIndex = randrange(5)
    phrases = ["Crack the Sky", "Let Your Hammer Fly", "Call Down the Lightning",
               "To Valhalla and Back", "Release Asgaard's Fury"]
    return phrases[phraseIndex-1]