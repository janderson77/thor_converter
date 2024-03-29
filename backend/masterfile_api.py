from openpyxl import load_workbook
from helpers import Employee, create_generic_import, collect_sheet_names
import re

def collect_hours(row, number, collecting=None):
    """
    Collects hours data from the appropriate cells
    Checks if there is data in the cell. If there is data, checks if it is already a float or not. If not, converts to float before returning.
    """
    if number == None:
            return None
    paydata = number
    if collecting == "bonus" or collecting == "commission" or collecting == "miles":
        if row[paydata] == None:
            if collecting == "miles":
                return None
            elif row[paydata+1] == None:
                return None
            paydata = paydata+1

        if type(row[paydata]) != float and type(row[paydata] != None):
            if type(row[paydata]) == str:
                return None
            return float(row[paydata])
        else:
            return row[paydata]
    elif type(row[paydata]) == None:
        return None
    if type(row[paydata]) == str and len(row[paydata]) < 1:
        return None
    elif type(row[paydata]) == float:
        return row[paydata]
    elif type(row[paydata]) == str and len(row[paydata]) > 0:
        if type(row[paydata]) == str:
            return None
        if re.search('[a-zA-Z]', row[paydata]):
            return None
        if row[paydata] == "#N/A":
            return None
        if "=" in row[paydata]:
            return None
        elif "." in row[paydata]:
            return float(row[paydata])
        else:
            h = int(row[paydata])
            return float(h)
    elif type(row[paydata]) == int:
        return float(row[paydata])
    else:
        return row[paydata]


def collect_employee_data(sheet, columns, s_name="None"):
    """Iterates over the sheet data, creating a new Employee class instance, filling it with data and appending it onto the data list"""
    data = []
    for row in sheet.iter_rows(min_row=sheet.min_row, max_row=sheet.max_row, values_only=True):
        tc = None
        if row[0] != None and type(row[0]) != str and row[0] <= 1:
            continue

        if row[0] != None and type(row[0]) == str and len(row[0]) > 2:
            if re.search('[a-zA-z]', row[0]):
                continue
            tc = Employee(int(row[0]))
            
        if row[0] != None and type(row[0]) != str:
            tc = Employee(id=row[0])
        
        if tc:
            reg1 = collect_hours(row, columns['reg1'])
            ot1 = collect_hours(row, columns['ot1'])
            salary = collect_hours(row, columns['salary'])
            bonus = collect_hours(row, columns['bonus'], "bonus")
            commission = collect_hours(row, columns['commission'], "commission")
            expenses = collect_hours(row, columns['expenses'])
            vacation = collect_hours(row, columns['vacation'])
            miles = collect_hours(row, columns['miles'], "miles")
            # Adds regular hours to time card, or sets the value to None
            if reg1 != None:
                tc.reg = round(reg1, 2)
            else:
                tc.reg = None

            # Adds overtime hours to time card, or sets the value to None
            if ot1 != None:
                tc.ot1 = round(ot1, 2)
            elif ot1 == None and tc.reg != None:
                tc.ot1 = 0
            else:
                tc.ot1 = None

            # Adds salary, bonus, commision and expense data, or leaves values as 0 or None
            if salary != None:
                tc.salary = round(salary, 2)

            if bonus != None: 
                tc.bonus = round(bonus, 2)
            elif bonus == None and tc.bonus != None:
                tc.bonus = None
            else:
                tc.bonus = None

            if commission != None:
                tc.commission = round(commission, 2)
            if expenses != None:
                if s_name == "DSD Managers":
                    tc.expenses = round(expenses, 2)
                    tc.adjustment = 9
                else:
                    tc.expenses = round(expenses, 2)
            if vacation != None:
                tc.vacation = round(vacation,2)
            if miles != None:
                tc.miles = round(miles,2)

            data.append(tc)
    return data


def find_data_row(sheet):
    """Finds the row where the column names live"""
    label_row = sheet.min_row
    for row in sheet.iter_rows(min_row=sheet.min_row, max_row=sheet.max_row, values_only=True):
        if "Emp #" not in row:
            label_row += 1
        elif row[0] == "Emp #":
            return label_row
        else:
            label_row += 1


def find_data_column(sheet, row):
    """Finds the column numbers for each required column"""
    columns = {
        "reg1": 6,
        "ot1": 7,
        "salary": None,
        "bonus": None,
        "commission": None,
        "expenses": 16,
        "vacation": None,
        "miles": None

    }

    data_row = None

    for row in sheet.iter_rows(min_row=row, max_row=row, values_only=True):
        data_row = row

    for i, v in enumerate(data_row[:24]):
        if v == None or type(v) == None:
            continue
        if "Reg Hrs" in v or "reg hrs" in v.lower():
            columns['reg1'] = i

        if "OT Hrs" in v or "ot hrs" in v.lower():
            columns['ot1'] = i

        if "salary" in v.lower():
            if "total" in v.lower():
                continue
            elif "pay" in v.lower():
                columns['salary'] = i
        if "bonus" in v.lower():
            if columns['bonus'] != None:
                continue
            if "rate" in v.lower():
                continue
            else:
                columns['bonus'] = i
        if "commission" in v.lower():
            if columns["commission"] != None:
                continue
            else:
                columns['commission'] = i
        if "Expenses-" in v or "Expenses -" in v:
            columns['expenses'] = i
        if "Vacation Pay" in v:
            columns['vacation'] = i
        if "miles" in v.lower() or "Milles" in v or "milles" in v.lower():
            columns["miles"] = i

    return columns


def convert_masterfile(input):
    """
    Converts all collected employee data from the Masterfile into importable excel files. DSD Managers is treated uniquely as they have cell phone adjustments.
    """
    wb = load_workbook(input, read_only=True)

    sheets = collect_sheet_names(wb,6)
    data = []
    special_data = []
    to_export = []

    for s in sheets:
        row = find_data_row(wb[s])
        columns = find_data_column(wb[s], row)
        if s == "DSD Managers":
            special_data.append([s, collect_employee_data(wb[s], columns, s)])
        else:
            data.append([s, collect_employee_data(wb[s], columns, s)])

    for d in data:
        res = create_generic_import(d, 1.162, "Papa Pita")
        to_export.append(res)

    if len(special_data) > 0:
        for d in special_data:
            gen_res = create_generic_import(d, 1.162, "Papa Pita")
            # adj_res = create_adjustment_import(d, 48, "Papa Pita")
            to_export.append(gen_res)
            # to_export.append(adj_res)

    return to_export
